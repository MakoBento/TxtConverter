"""
TxtConverter - 各種テキストファイル・Excel(.xlsx/.xlsm/.xls)をテキストに変換するツール
"""

import os
import sys
import threading
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from pathlib import Path

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from tkinterdnd2 import DND_FILES, TkinterDnD


# ---------------------------------------------------------------------------
# 対応拡張子の定義
# ---------------------------------------------------------------------------

# Excel系: 複雑な変換（取り消し線判定など）を行う
EXCEL_EXTENSIONS = (".xlsx", ".xlsm", ".xls")

# テキスト系: 内容をそのままtxtとして出力する
TEXT_EXTENSIONS = (
    ".cs", ".md", ".txt", ".sql", ".py", ".js",
    ".html", ".htm", ".ts", ".tsx", ".css", ".vue", ".json", ".xml"
)

# 全対応拡張子
ALL_EXTENSIONS = EXCEL_EXTENSIONS + TEXT_EXTENSIONS


# ---------------------------------------------------------------------------
# Excel変換ロジック
# ---------------------------------------------------------------------------

def _extract_cell_text(cell) -> str:
    """セルからテキストを抽出する。取り消し線が付いた文字は<deleted>タグで囲む。"""
    value = cell.value
    if value is None:
        return ""

    # --- Rich Text (セル内で部分的に書式が異なる場合) ---
    if isinstance(value, CellRichText):
        parts: list[str] = []
        for element in value:
            if isinstance(element, str):
                # プレーンテキスト部分: セル全体のフォントで判定
                if cell.font and cell.font.strike:
                    parts.append(f"<deleted>{element}</deleted>")
                else:
                    parts.append(element)
            elif isinstance(element, TextBlock):
                # 書式付きラン: ラン固有のフォントで判定
                text = element.text or ""
                if element.font and element.font.strike:
                    parts.append(f"<deleted>{text}</deleted>")
                else:
                    parts.append(text)
        return "".join(parts)

    # --- 通常テキスト ---
    if cell.font and cell.font.strike:
        return f"<deleted>{value}</deleted>"

    return str(value)


def convert_workbook(wb, sheet_names=None) -> str:
    """ワークブックオブジェクトをテキストに変換する。"""
    lines: list[str] = []
    targets = sheet_names or wb.sheetnames

    for sname in targets:
        ws = wb[sname]
        lines.append(f"--- Sheet: {sname} ---")

        for row in ws.iter_rows():
            cells_text: list[str] = []
            for cell in row:
                cells_text.append(_extract_cell_text(cell))
            # 末尾の空セルを削除して出力を軽量化
            while cells_text and cells_text[-1] == "":
                cells_text.pop()
            lines.append("\t".join(cells_text))

        lines.append("")  # シート間に空行

    return "\n".join(lines)


def _convert_xls(src_path: str) -> str:
    """
    .xls ファイルを変換する。
    xlrd がインストールされていれば使用する。
    xlrd は書式の取り消し線取得が困難なため、全テキストをそのまま出力する。
    """
    try:
        import xlrd
    except ImportError:
        raise RuntimeError(
            ".xls ファイルの読み込みには xlrd が必要です。"
            "pip install xlrd でインストールしてください。"
        )

    book = xlrd.open_workbook(src_path, formatting_info=False)
    lines: list[str] = []

    for sheet in book.sheets():
        lines.append(f"--- Sheet: {sheet.name} ---")
        for rx in range(sheet.nrows):
            row_vals: list[str] = []
            for cx in range(sheet.ncols):
                cell = sheet.cell(rx, cx)
                row_vals.append(str(cell.value) if cell.value is not None else "")
            while row_vals and row_vals[-1] == "":
                row_vals.pop()
            lines.append("\t".join(row_vals))
        lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# テキスト系変換ロジック
# ---------------------------------------------------------------------------

def _read_text_file(src_path: str) -> str:
    """
    テキストファイルを読み込む。
    UTF-8 → CP932(Shift-JIS) → UTF-8(errors=replace) の順でフォールバックする。
    """
    encodings = ["utf-8", "cp932"]
    for enc in encodings:
        try:
            return Path(src_path).read_text(encoding=enc)
        except (UnicodeDecodeError, LookupError):
            continue
    # 最終フォールバック: 読めない文字を置換
    return Path(src_path).read_text(encoding="utf-8", errors="replace")


# ---------------------------------------------------------------------------
# 共通変換エントリポイント
# ---------------------------------------------------------------------------

def convert_file(src_path: str, dst_dir: str, log_func) -> bool:
    """1ファイルを変換して出力先に保存する。成功時 True。"""
    src = Path(src_path)
    ext = src.suffix.lower()
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")

    try:
        if ext in (".xlsx", ".xlsm"):
            # .xlsm はマクロ有効ブック。openpyxl はマクロを無視して読み込める
            # rich_text=True で CellRichText を取得可能にする
            wb = openpyxl.load_workbook(
                str(src), read_only=True, data_only=True, rich_text=True
            )
            text = convert_workbook(wb)
            wb.close()

        elif ext == ".xls":
            # .xls は openpyxl 非対応。xlrd で読み取り後テキスト化する
            text = _convert_xls(str(src))

        elif ext in TEXT_EXTENSIONS:
            # テキスト系: 内容をそのままtxtとして出力する
            text = _read_text_file(str(src))

        else:
            log_func(f"[スキップ] {src.name}: 非対応の拡張子")
            return False

        # 出力ファイル名: 元のファイル名（拡張子含む）_タイムスタンプ.txt
        out_path = Path(dst_dir) / f"{src.name}_{timestamp}.txt"
        out_path.write_text(text, encoding="utf-8")
        log_func(f"[完了] {src.name} -> {out_path.name}")
        return True

    except Exception as e:
        log_func(f"[エラー] {src.name}: {e}")
        return False


def collect_target_files(folder: str) -> list[str]:
    """フォルダ内（サブフォルダ含む）の全対応拡張子ファイルを再帰的に収集して返す。"""
    result = []
    for root, _dirs, files in os.walk(folder):
        for f in files:
            if Path(f).suffix.lower() in ALL_EXTENSIONS:
                result.append(os.path.join(root, f))
    return result


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------

class App(TkinterDnD.Tk):
    def __init__(self):
        super().__init__()
        self.title("TxtConverter")
        self.geometry("720x520")
        self.resizable(True, True)
        self._build_ui()

    # ----- UI構築 -----
    def _build_ui(self):
        pad = {"padx": 8, "pady": 4}

        # --- 使い方説明 ---
        help_text = (
            "ファイルまたはフォルダを指定して.txtファイルに変換します（サブフォルダも対象）\n"
            "excel系のファイルは取消線ありのテキストを<deleted>タグで囲います\n"
            "対応: .xlsx/.xlsm/.xls/.cs/.md/.txt/.sql/.py/.js/.html/.htm/.ts/.tsx/.css/.vue/.json/.xml\n"
            "出力: 指定フォルダに「元ファイル名_yyyymmddhhmmss.txt」で保存（例: main.py_20260322120000.txt）"
        )
        tk.Label(
            self, text=help_text, justify="left",
            anchor="w", relief="groove", padx=8, pady=6
        ).pack(fill="x", padx=8, pady=(8, 2))

        # --- 入力パス（ファイル・フォルダ自動判定） ---
        src_frame = tk.LabelFrame(self, text="対象ファイル / フォルダ")
        src_frame.pack(fill="x", **pad)

        self.src_var = tk.StringVar()
        self.src_entry = tk.Entry(src_frame, textvariable=self.src_var)
        self.src_entry.pack(side="left", fill="x", expand=True, padx=(8, 0), pady=4)
        self.src_entry.drop_target_register(DND_FILES)
        self.src_entry.dnd_bind("<<Drop>>", self._on_drop)
        # フォルダ参照ボタン
        self.src_btn_folder = tk.Button(
            src_frame, text="フォルダ参照", command=self._browse_src_folder
        )
        self.src_btn_folder.pack(side="right", padx=(0, 4), pady=4)
        # ファイル参照ボタン
        self.src_btn_file = tk.Button(
            src_frame, text="ファイル参照", command=self._browse_src_file
        )
        self.src_btn_file.pack(side="right", padx=(8, 0), pady=4)

        # --- 出力先 ---
        dst_frame = tk.LabelFrame(self, text="出力先フォルダ")
        dst_frame.pack(fill="x", **pad)

        self.dst_var = tk.StringVar()
        self.dst_entry = tk.Entry(dst_frame, textvariable=self.dst_var)
        self.dst_entry.pack(side="left", fill="x", expand=True, padx=(8, 0), pady=4)
        self.dst_entry.drop_target_register(DND_FILES)
        self.dst_entry.dnd_bind("<<Drop>>", self._on_drop_dst)
        tk.Button(dst_frame, text="参照...", command=self._browse_dst
                  ).pack(side="right", padx=8, pady=4)

        # --- 実行ボタン ---
        self.run_btn = tk.Button(
            self, text="変換実行", width=20, command=self._on_run
        )
        self.run_btn.pack(pady=8)

        # --- ログ ---
        log_frame = tk.LabelFrame(self, text="ログ")
        log_frame.pack(fill="both", expand=True, **pad)

        self.log_area = scrolledtext.ScrolledText(
            log_frame, height=12, state="disabled", wrap="word"
        )
        self.log_area.pack(fill="both", expand=True, padx=4, pady=4)

    # ----- イベントハンドラ -----
    @staticmethod
    def _parse_drop_paths(raw: str) -> list[str]:
        """ドロップイベントのデータからパスのリストを解析する。"""
        raw = raw.strip()
        paths = []
        i = 0
        while i < len(raw):
            if raw[i] == "{":
                end = raw.index("}", i)
                paths.append(raw[i + 1:end])
                i = end + 2  # '}' の次のスペースをスキップ
            else:
                space_idx = raw.find(" ", i)
                if space_idx == -1:
                    paths.append(raw[i:])
                    break
                else:
                    paths.append(raw[i:space_idx])
                    i = space_idx + 1
        return paths

    def _on_drop(self, event):
        """ドラッグ&ドロップでファイル/フォルダパスを設定する。"""
        paths = self._parse_drop_paths(event.data)
        if paths:
            self.src_var.set(paths[0])

    def _on_drop_dst(self, event):
        """出力先フォルダへのドラッグ&ドロップ。ディレクトリのみ許可する。"""
        paths = self._parse_drop_paths(event.data)
        if not paths:
            return
        path = paths[0]
        if os.path.isdir(path):
            self.dst_var.set(path)
        else:
            messagebox.showwarning("入力エラー", "フォルダのみドロップできます。")

    def _browse_src_file(self):
        """ファイル選択ダイアログを開く。"""
        all_exts = " ".join(f"*{e}" for e in ALL_EXTENSIONS)
        excel_exts = " ".join(f"*{e}" for e in EXCEL_EXTENSIONS)
        text_exts = " ".join(f"*{e}" for e in TEXT_EXTENSIONS)
        path = filedialog.askopenfilename(
            title="ファイルを選択",
            filetypes=[
                ("全対応ファイル", all_exts),
                ("Excel ファイル", excel_exts),
                ("テキスト系ファイル", text_exts),
                ("すべてのファイル", "*.*"),
            ]
        )
        if path:
            self.src_var.set(path)

    def _browse_src_folder(self):
        """フォルダ選択ダイアログを開く。"""
        path = filedialog.askdirectory(title="フォルダを選択")
        if path:
            self.src_var.set(path)

    def _browse_dst(self):
        path = filedialog.askdirectory(title="出力先フォルダを選択")
        if path:
            self.dst_var.set(path)

    def _log(self, msg: str):
        """スレッドセーフにログを追加する。"""
        def _append():
            self.log_area.configure(state="normal")
            self.log_area.insert("end", msg + "\n")
            self.log_area.see("end")
            self.log_area.configure(state="disabled")
        self.after(0, _append)

    def _on_run(self):
        src = self.src_var.get().strip()
        dst = self.dst_var.get().strip()

        if not src:
            messagebox.showwarning("入力エラー", "対象ファイル/フォルダを指定してください。")
            return
        if not dst:
            messagebox.showwarning("入力エラー", "出力先フォルダを指定してください。")
            return
        if not os.path.exists(dst):
            messagebox.showwarning("入力エラー", "出力先フォルダが存在しません。")
            return

        # パスの種類を自動判定してファイルリストを作成
        if os.path.isfile(src):
            # 単一ファイル
            ext = Path(src).suffix.lower()
            if ext not in ALL_EXTENSIONS:
                exts_str = "、".join(ALL_EXTENSIONS)
                messagebox.showwarning(
                    "入力エラー",
                    f"対応していないファイル形式です。\n対応拡張子: {exts_str}"
                )
                return
            files = [src]
        elif os.path.isdir(src):
            # フォルダ一括
            files = collect_target_files(src)
            if not files:
                messagebox.showwarning("入力エラー", "フォルダ内に対応ファイルが見つかりません。")
                return
        else:
            messagebox.showwarning("入力エラー", "指定されたファイル/フォルダが見つかりません。")
            return

        # UIロックしてバックグラウンド実行
        self.run_btn.configure(state="disabled")
        self.src_btn_file.configure(state="disabled")
        self.src_btn_folder.configure(state="disabled")
        self._log(f"--- 変換開始 ({len(files)} ファイル) ---")

        thread = threading.Thread(target=self._run_convert, args=(files, dst), daemon=True)
        thread.start()

    def _run_convert(self, files: list[str], dst: str):
        ok = 0
        ng = 0
        for f in files:
            self._log(f"処理中: {os.path.basename(f)}")
            if convert_file(f, dst, self._log):
                ok += 1
            else:
                ng += 1

        self._log(f"--- 変換完了: 成功={ok}, 失敗/スキップ={ng} ---\n")
        self.after(0, lambda: self.run_btn.configure(state="normal"))
        self.after(0, lambda: self.src_btn_file.configure(state="normal"))
        self.after(0, lambda: self.src_btn_folder.configure(state="normal"))


# ---------------------------------------------------------------------------
# エントリポイント
# ---------------------------------------------------------------------------

def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
