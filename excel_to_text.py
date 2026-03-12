"""
ExcelToText - Excel(.xlsx/.xlsm/.xls)ファイルを取り消し線を除外してテキストに変換するツール
"""

import os
import sys
import threading
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from pathlib import Path

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from tkinterdnd2 import DND_FILES, TkinterDnD


# ---------------------------------------------------------------------------
# 変換ロジック
# ---------------------------------------------------------------------------

def _extract_cell_text(cell) -> str:
    """セルからテキストを抽出する。取り消し線が付いた文字は<deleted>タグで囲む。"""
    value = cell.value
    if value is None:
        return ""

    # --- Rich Text (セル内で部分的に書式が異なる場合) ---
    if isinstance(value, CellRichText):
        parts: list[str] = []
        for element in value:
            if isinstance(element, str):
                # プレーンテキスト部分: セル全体のフォントで判定
                if cell.font and cell.font.strike:
                    parts.append(f"<deleted>{element}</deleted>")
                else:
                    parts.append(element)
            elif isinstance(element, TextBlock):
                # 書式付きラン: ラン固有のフォントで判定
                text = element.text or ""
                if element.font and element.font.strike:
                    parts.append(f"<deleted>{text}</deleted>")
                else:
                    parts.append(text)
        return "".join(parts)

    # --- 通常テキスト ---
    if cell.font and cell.font.strike:
        return f"<deleted>{value}</deleted>"

    return str(value)


def convert_workbook(wb, sheet_names=None) -> str:
    """ワークブックオブジェクトをテキストに変換する。"""
    lines: list[str] = []
    targets = sheet_names or wb.sheetnames

    for sname in targets:
        ws = wb[sname]
        lines.append(f"--- Sheet: {sname} ---")

        for row in ws.iter_rows():
            cells_text: list[str] = []
            for cell in row:
                cells_text.append(_extract_cell_text(cell))
            # 末尾の空セルを削除して出力を軽量化
            while cells_text and cells_text[-1] == "":
                cells_text.pop()
            lines.append("\t".join(cells_text))

        lines.append("")  # シート間に空行

    return "\n".join(lines)


def convert_file(src_path: str, dst_dir: str, log_func) -> bool:
    """1ファイルを変換して出力先に保存する。成功時 True。"""
    src = Path(src_path)
    ext = src.suffix.lower()
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")

    try:
        if ext in (".xlsx", ".xlsm"):
            # .xlsm はマクロ有効ブック。openpyxl はマクロを無視して読み込める
            # rich_text=True で CellRichText を取得可能にする
            wb = openpyxl.load_workbook(
                str(src), read_only=True, data_only=True, rich_text=True
            )
        elif ext == ".xls":
            # .xls は openpyxl 非対応。xlrd で読み取り後テキスト化する
            text = _convert_xls(str(src))
            out_path = Path(dst_dir) / f"{src.stem}_{timestamp}.txt"
            out_path.write_text(text, encoding="utf-8")
            log_func(f"[完了] {src.name} -> {out_path.name}")
            return True
        else:
            log_func(f"[スキップ] {src.name}: 非対応の拡張子")
            return False

        text = convert_workbook(wb)
        wb.close()

        out_path = Path(dst_dir) / f"{src.stem}_{timestamp}.txt"
        out_path.write_text(text, encoding="utf-8")
        log_func(f"[完了] {src.name} -> {out_path.name}")
        return True

    except Exception as e:
        log_func(f"[エラー] {src.name}: {e}")
        return False


def _convert_xls(src_path: str) -> str:
    """
    .xls ファイルを変換する。
    xlrd がインストールされていれば使用する。
    xlrd は書式の取り消し線取得が困難なため、全テキストをそのまま出力する。
    """
    try:
        import xlrd
    except ImportError:
        raise RuntimeError(
            ".xls ファイルの読み込みには xlrd が必要です。"
            "pip install xlrd でインストールしてください。"
        )

    book = xlrd.open_workbook(src_path, formatting_info=False)
    lines: list[str] = []

    for sheet in book.sheets():
        lines.append(f"--- Sheet: {sheet.name} ---")
        for rx in range(sheet.nrows):
            row_vals: list[str] = []
            for cx in range(sheet.ncols):
                cell = sheet.cell(rx, cx)
                row_vals.append(str(cell.value) if cell.value is not None else "")
            while row_vals and row_vals[-1] == "":
                row_vals.pop()
            lines.append("\t".join(row_vals))
        lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------

class App(TkinterDnD.Tk):
    def __init__(self):
        super().__init__()
        self.title("ExcelToText")
        self.geometry("720x520")
        self.resizable(True, True)
        self._build_ui()

    # ----- UI構築 -----
    def _build_ui(self):
        pad = {"padx": 8, "pady": 4}

        # --- モード選択 ---
        mode_frame = tk.LabelFrame(self, text="変換モード")
        mode_frame.pack(fill="x", **pad)

        self.mode_var = tk.StringVar(value="file")
        tk.Radiobutton(
            mode_frame, text="単一ファイルモード", variable=self.mode_var,
            value="file", command=self._on_mode_change
        ).pack(side="left", padx=12, pady=4)
        tk.Radiobutton(
            mode_frame, text="フォルダ一括モード", variable=self.mode_var,
            value="folder", command=self._on_mode_change
        ).pack(side="left", padx=12, pady=4)

        # --- 入力パス ---
        src_frame = tk.LabelFrame(self, text="対象ファイル / フォルダ")
        src_frame.pack(fill="x", **pad)

        self.src_var = tk.StringVar()
        self.src_entry = tk.Entry(src_frame, textvariable=self.src_var)
        self.src_entry.pack(side="left", fill="x", expand=True, padx=(8, 0), pady=4)
        self.src_entry.drop_target_register(DND_FILES)
        self.src_entry.dnd_bind("<<Drop>>", self._on_drop)
        self.src_btn = tk.Button(src_frame, text="参照...", command=self._browse_src)
        self.src_btn.pack(side="right", padx=8, pady=4)

        # --- 出力先 ---
        dst_frame = tk.LabelFrame(self, text="出力先フォルダ")
        dst_frame.pack(fill="x", **pad)

        self.dst_var = tk.StringVar()
        self.dst_entry = tk.Entry(dst_frame, textvariable=self.dst_var)
        self.dst_entry.pack(side="left", fill="x", expand=True, padx=(8, 0), pady=4)
        self.dst_entry.drop_target_register(DND_FILES)
        self.dst_entry.dnd_bind("<<Drop>>", self._on_drop_dst)
        tk.Button(dst_frame, text="参照...", command=self._browse_dst
                  ).pack(side="right", padx=8, pady=4)

        # --- 実行ボタン ---
        self.run_btn = tk.Button(
            self, text="変換実行", width=20, command=self._on_run
        )
        self.run_btn.pack(pady=8)

        # --- ログ ---
        log_frame = tk.LabelFrame(self, text="ログ")
        log_frame.pack(fill="both", expand=True, **pad)

        self.log_area = scrolledtext.ScrolledText(
            log_frame, height=12, state="disabled", wrap="word"
        )
        self.log_area.pack(fill="both", expand=True, padx=4, pady=4)

    # ----- イベントハンドラ -----
    def _on_mode_change(self):
        self.src_var.set("")

    @staticmethod
    def _parse_drop_paths(raw: str) -> list[str]:
        """ドロップイベントのデータからパスのリストを解析する。"""
        raw = raw.strip()
        paths = []
        i = 0
        while i < len(raw):
            if raw[i] == "{":
                end = raw.index("}", i)
                paths.append(raw[i + 1:end])
                i = end + 2  # '}' の次のスペースをスキップ
            else:
                space_idx = raw.find(" ", i)
                if space_idx == -1:
                    paths.append(raw[i:])
                    break
                else:
                    paths.append(raw[i:space_idx])
                    i = space_idx + 1
        return paths

    def _on_drop(self, event):
        """ドラッグ&ドロップでファイル/フォルダパスを設定する。"""
        paths = self._parse_drop_paths(event.data)
        if paths:
            self.src_var.set(paths[0])

    def _on_drop_dst(self, event):
        """出力先フォルダへのドラッグ&ドロップ。ディレクトリのみ許可する。"""
        paths = self._parse_drop_paths(event.data)
        if not paths:
            return
        path = paths[0]
        if os.path.isdir(path):
            self.dst_var.set(path)
        else:
            messagebox.showwarning("入力エラー", "フォルダのみドロップできます。")

    def _browse_src(self):
        if self.mode_var.get() == "file":
            path = filedialog.askopenfilename(
                title="Excelファイルを選択",
                filetypes=[("Excel ファイル", "*.xlsx *.xlsm *.xls")]
            )
        else:
            path = filedialog.askdirectory(title="フォルダを選択")
        if path:
            self.src_var.set(path)

    def _browse_dst(self):
        path = filedialog.askdirectory(title="出力先フォルダを選択")
        if path:
            self.dst_var.set(path)

    def _log(self, msg: str):
        """スレッドセーフにログを追加する。"""
        def _append():
            self.log_area.configure(state="normal")
            self.log_area.insert("end", msg + "\n")
            self.log_area.see("end")
            self.log_area.configure(state="disabled")
        self.after(0, _append)

    def _on_run(self):
        src = self.src_var.get().strip()
        dst = self.dst_var.get().strip()

        if not src:
            messagebox.showwarning("入力エラー", "対象ファイル/フォルダを指定してください。")
            return
        if not dst:
            messagebox.showwarning("入力エラー", "出力先フォルダを指定してください。")
            return
        if not os.path.exists(dst):
            messagebox.showwarning("入力エラー", "出力先フォルダが存在しません。")
            return

        # ファイルリスト作成
        if self.mode_var.get() == "file":
            if not os.path.isfile(src):
                messagebox.showwarning("入力エラー", "指定されたファイルが見つかりません。")
                return
            if not src.lower().endswith((".xlsx", ".xlsm", ".xls")):
                messagebox.showwarning("入力エラー", "対応していないファイル形式です。\n.xlsx、.xlsm または .xls ファイルを指定してください。")
                return
            files = [src]
        else:
            if not os.path.isdir(src):
                messagebox.showwarning("入力エラー", "指定されたフォルダが見つかりません。")
                return
            files = [
                os.path.join(src, f) for f in os.listdir(src)
                if f.lower().endswith((".xlsx", ".xlsm", ".xls"))
            ]
            if not files:
                messagebox.showwarning("入力エラー", "フォルダ内にExcelファイルが見つかりません。")
                return

        # UIロックしてバックグラウンド実行
        self.run_btn.configure(state="disabled")
        self.src_btn.configure(state="disabled")
        self._log(f"--- 変換開始 ({len(files)} ファイル) ---")

        thread = threading.Thread(target=self._run_convert, args=(files, dst), daemon=True)
        thread.start()

    def _run_convert(self, files: list[str], dst: str):
        ok = 0
        ng = 0
        for f in files:
            self._log(f"処理中: {os.path.basename(f)}")
            if convert_file(f, dst, self._log):
                ok += 1
            else:
                ng += 1

        self._log(f"--- 変換完了: 成功={ok}, 失敗/スキップ={ng} ---\n")
        self.after(0, lambda: self.run_btn.configure(state="normal"))
        self.after(0, lambda: self.src_btn.configure(state="normal"))


# ---------------------------------------------------------------------------
# エントリポイント
# ---------------------------------------------------------------------------

def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
